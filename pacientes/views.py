import datetime

from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from django.utils.dateparse import parse_date

from accounts.models import Bitacora, Usuario
from accounts.views import es_administrador

from .forms import (
    AdjuntarImagenesForm,
    AdjuntarInformeForm,
    AgendarCitaForm,
    CompletarDatosPacienteForm,
    CrearTipoEstudioForm,
    GenerarOrdenForm,
    ProcesarTicketForm,
    RegistrarTicketForm,
)
from .horarios import (
    DIAS_SEMANA,
    LIMITE_DIAS_ADELANTE,
    PASO_MINUTOS,
    en_el_pasado,
    fuera_de_ventana,
    horarios_disponibles,
    inicio_semana,
    rango_ocupado_por,
    se_cruzan,
)
from .models import (
    Cita,
    ImagenEstudio,
    Notificacion,
    OrdenTrabajo,
    Paciente,
    ReporteDiario,
    Ticket,
    TipoEstudio,
)


# Cuántas citas de emergencia (agendadas encima de otra ya existente) se
# permiten como máximo por día, para no saturar a la radióloga.
MAXIMO_EMERGENCIAS_POR_DIA = 5


def es_recepcionista(user):
    return user.is_authenticated and user.rol == Usuario.ROL_RECEPCIONISTA


def es_tecnico(user):
    return user.is_authenticated and user.rol == Usuario.ROL_TECNICO_IMAGENES


def es_radiologo(user):
    return user.is_authenticated and user.rol == Usuario.ROL_MEDICO_RADIOLOGO


def es_administrador_financiero(user):
    return user.is_authenticated and (user.is_superuser or user.rol == Usuario.ROL_ADMINISTRADOR_FINANCIERO)


def puede_ver_reportes_diarios(user):
    return es_recepcionista(user) or es_administrador_financiero(user) or es_administrador(user)


def puede_descargar_reportes_diarios(user):
    return es_administrador_financiero(user) or es_administrador(user)


def _notificar_cita_asignada(cita):
    """El radiólogo elegido al agendar la cita recibe una nueva solicitud
    para revisar/confirmar. Si la recepcionista la marcó como emergencia
    (agendada encima de otra cita ya existente), el mensaje lo deja claro
    para que la radióloga sepa que se le está pidiendo hacer un espacio."""
    prefijo = '🚨 EMERGENCIA — ' if cita.es_emergencia_forzada else ''
    mensaje = (
        f'{prefijo}Nueva cita asignada: {cita.tipo_estudio} para {cita.paciente.nombre} '
        f'{cita.paciente.apellido} el {cita.fecha_sugerida or cita.fecha} a las '
        f'{cita.hora_sugerida or cita.hora}.'
    )
    if cita.es_emergencia_forzada:
        mensaje += ' Este horario ya tenía otra cita asignada: se agendó igual por ser una emergencia.'
    Notificacion.notificar(
        destinatario=cita.radiologo,
        tipo=Notificacion.TIPO_CITA_ASIGNADA,
        mensaje=mensaje,
        cita=cita,
        url=reverse('solicitudes_pendientes'),
    )


def _notificar_cita_confirmada(cita):
    """El radiólogo confirmó fecha/hora de la solicitud: se avisa a quien
    la creó (recepción) para que pueda comunicárselo al paciente."""
    Notificacion.notificar(
        destinatario=cita.creada_por,
        tipo=Notificacion.TIPO_CITA_CONFIRMADA,
        mensaje=(
            f'Cita confirmada: {cita.tipo_estudio} de {cita.paciente.nombre} '
            f'{cita.paciente.apellido} el {cita.fecha} a las {cita.hora}.'
        ),
        cita=cita,
        url=reverse(f'calendario_{cita.convenio}'),
    )


def _notificar_cita_rechazada(cita):
    """El radiólogo rechazó la solicitud: se avisa a quien la creó."""
    Notificacion.notificar(
        destinatario=cita.creada_por,
        tipo=Notificacion.TIPO_CITA_RECHAZADA,
        mensaje=(
            f'Cita rechazada: {cita.tipo_estudio} de {cita.paciente.nombre} '
            f'{cita.paciente.apellido}. Motivo: {cita.motivo_rechazo or "—"}'
        ),
        cita=cita,
        url=reverse(f'calendario_{cita.convenio}'),
    )


def _notificar_orden_pendiente(cita):
    """Aviso para el equipo de técnicos: hay una orden de trabajo nueva
    esperando que le tomen las imágenes."""
    tecnicos = Usuario.objects.filter(rol=Usuario.ROL_TECNICO_IMAGENES, is_active=True)
    Notificacion.notificar_a_varios(
        usuarios=tecnicos,
        tipo=Notificacion.TIPO_ORDEN_PENDIENTE,
        mensaje=(
            f'Nueva orden de trabajo: {cita.tipo_estudio} para {cita.paciente.nombre} '
            f'{cita.paciente.apellido}.'
        ),
        cita=cita,
        url=reverse('ordenes_pendientes'),
    )


def _notificar_estudio_listo_para_informar(cita):
    """Cuando el técnico termina de subir las imágenes, se avisa al
    radiólogo asignado (o a todo el equipo de radiología si la cita no
    tenía uno asignado, como los tickets de emergencia)."""
    if cita.radiologo_id:
        radiologos = [cita.radiologo]
    else:
        radiologos = Usuario.objects.filter(rol=Usuario.ROL_MEDICO_RADIOLOGO, is_active=True)
    Notificacion.notificar_a_varios(
        usuarios=radiologos,
        tipo=Notificacion.TIPO_ESTUDIO_LISTO_INFORMAR,
        mensaje=(
            f'Estudio listo para informar: {cita.tipo_estudio} de {cita.paciente.nombre} '
            f'{cita.paciente.apellido}.'
        ),
        cita=cita,
        url=reverse('citas_procesadas'),
    )


def _notificar_estudio_completado(cita):
    """Cuando el radiólogo termina el informe (diagnóstico), se avisa a
    todo el equipo de recepción para que puedan entregar el resultado."""
    try:
        url = reverse(f'procesar_citas_{cita.convenio}')
    except NoReverseMatch:
        # Todavía no existe una pantalla de "procesar citas" para este
        # convenio (por ahora solo COEX la tiene); igual se notifica, solo
        # que el enlace de la notificación cae al panel principal.
        url = reverse('dashboard')
    recepcionistas = Usuario.objects.filter(rol=Usuario.ROL_RECEPCIONISTA, is_active=True)
    Notificacion.notificar_a_varios(
        usuarios=recepcionistas,
        tipo=Notificacion.TIPO_ESTUDIO_COMPLETADO,
        mensaje=(
            f'Estudio completado: {cita.tipo_estudio} de {cita.paciente.nombre} {cita.paciente.apellido}. '
            'Informe y diagnóstico listos.'
        ),
        cita=cita,
        url=url,
    )


def _notificar_reporte_enviado(reporte, enviado_por):
    """Cuando la recepcionista envía el reporte diario, se avisa a todo el
    equipo de administración financiera para que puedan revisarlo."""
    destinatarios = Usuario.objects.filter(rol=Usuario.ROL_ADMINISTRADOR_FINANCIERO, is_active=True)
    Notificacion.notificar_a_varios(
        usuarios=destinatarios,
        tipo=Notificacion.TIPO_REPORTE_ENVIADO,
        mensaje=(
            f'Nuevo reporte diario enviado: {reporte.get_convenio_display()} del {reporte.fecha}, '
            f'por {enviado_por.get_full_name() or enviado_por.username}.'
        ),
        url=reverse('ver_reporte_diario', args=[reporte.convenio, reporte.fecha.strftime("%Y-%m-%d")]),
    )


CAMPOS_DATOS_PACIENTE = ('nombre', 'apellido', 'sexo', 'telefono', 'fecha_nacimiento')


def obtener_o_actualizar_paciente(cd):
    """Reutiliza el paciente si el DPI ya existe (evita duplicar el registro)
    y sincroniza sus datos con lo capturado en el formulario, por si el
    recepcionista corrigió algo (ej. un teléfono desactualizado).

    El carné de afiliación IGSS, el sexo, el teléfono y la fecha de
    nacimiento son opcionales (ej. registro apurado en una emergencia) y se
    identifican siempre por el DPI del paciente: si vienen vacíos no se
    borra el valor que ya tuviera registrado de una visita anterior."""
    carnet_igss = cd.get('carnet_igss') or None
    paciente, creado = Paciente.objects.get_or_create(
        dpi=cd['dpi'],
        defaults={**{campo: cd[campo] for campo in CAMPOS_DATOS_PACIENTE}, 'carnet_igss': carnet_igss},
    )
    if not creado:
        cambiados = []
        for campo in CAMPOS_DATOS_PACIENTE:
            valor_nuevo = cd[campo]
            si_opcional_y_vacio = campo in Paciente.CAMPOS_OPCIONALES and not valor_nuevo
            if si_opcional_y_vacio:
                continue
            if getattr(paciente, campo) != valor_nuevo:
                setattr(paciente, campo, valor_nuevo)
                cambiados.append(campo)
        if carnet_igss and paciente.carnet_igss != carnet_igss:
            paciente.carnet_igss = carnet_igss
            cambiados.append('carnet_igss')
        if cambiados:
            paciente.save(update_fields=cambiados)
    _notificar_datos_pendientes_si_corresponde(paciente)
    return paciente


def _notificar_datos_pendientes_si_corresponde(paciente):
    """Si el paciente quedó con datos opcionales sin llenar (sexo, teléfono
    o fecha de nacimiento), avisa a las recepcionistas activas. No duplica
    el aviso si ya hay uno sin leer para este mismo paciente — por eso se
    puede llamar cada vez que se agenda una cita o se registra un ticket
    sin que se acumulen notificaciones repetidas."""
    campos = paciente.campos_pendientes()
    if not campos:
        return
    url = reverse('completar_datos_paciente', args=[paciente.id])
    ya_avisado = Notificacion.objects.filter(
        tipo=Notificacion.TIPO_DATOS_PACIENTE_PENDIENTES, url=url, leida=False,
    ).exists()
    if ya_avisado:
        return
    recepcionistas = Usuario.objects.filter(rol=Usuario.ROL_RECEPCIONISTA, is_active=True)
    Notificacion.notificar_a_varios(
        usuarios=recepcionistas,
        tipo=Notificacion.TIPO_DATOS_PACIENTE_PENDIENTES,
        mensaje=(
            f'{paciente.nombre} {paciente.apellido} (DPI {paciente.dpi}) tiene datos '
            f'pendientes: {", ".join(campos)}.'
        ),
        url=url,
    )


@login_required
@user_passes_test(es_recepcionista)
def completar_datos_paciente(request, paciente_id):
    """Pantalla a la que llega la recepcionista al hacer clic en la
    notificación de "datos pendientes": deja llenar sexo, teléfono y/o
    fecha de nacimiento sin tener que pasar de nuevo por agendar una cita."""
    paciente = get_object_or_404(Paciente, id=paciente_id)

    if request.method == 'POST':
        form = CompletarDatosPacienteForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            cambiados = [
                campo for campo in ('sexo', 'telefono', 'fecha_nacimiento')
                if cd[campo] and getattr(paciente, campo) != cd[campo]
            ]
            for campo in cambiados:
                setattr(paciente, campo, cd[campo])
            if cambiados:
                paciente.save(update_fields=cambiados)
                messages.success(
                    request, f'Datos de {paciente.nombre} {paciente.apellido} actualizados.'
                )
            else:
                messages.info(request, 'No se cargó ningún dato nuevo.')

            if paciente.campos_pendientes():
                return redirect('completar_datos_paciente', paciente_id=paciente.id)

            # Recién ahora quedó completo: recién ahora se apaga el aviso
            # (para cualquier recepcionista, no solo quien lo llenó). Hasta
            # este punto la notificación se queda, aunque ya se haya
            # abierto/leído esta pantalla.
            Notificacion.objects.filter(
                tipo=Notificacion.TIPO_DATOS_PACIENTE_PENDIENTES,
                url=reverse('completar_datos_paciente', args=[paciente.id]),
                leida=False,
            ).update(leida=True)
            return redirect('dashboard')
    else:
        form = CompletarDatosPacienteForm(initial={
            'sexo': paciente.sexo,
            'telefono': paciente.telefono,
            'fecha_nacimiento': paciente.fecha_nacimiento,
        })

    return render(request, 'pacientes/completar_datos_paciente.html', {
        'form': form,
        'paciente': paciente,
        'pendientes': paciente.campos_pendientes(),
    })


@login_required
@user_passes_test(es_recepcionista)
def buscar_paciente_por_dpi(request):
    """Usado por el formulario de agendar cita / registrar ticket para
    autocompletar los datos si el paciente ya está registrado, en vez de
    hacer que el recepcionista los vuelva a escribir."""
    dpi = (request.GET.get('dpi') or '').strip()
    paciente = Paciente.objects.filter(dpi=dpi).first() if dpi else None
    if not paciente:
        return JsonResponse({'encontrado': False})
    return JsonResponse({
        'encontrado': True,
        'nombre': paciente.nombre,
        'apellido': paciente.apellido,
        'sexo': paciente.sexo,
        'telefono': paciente.telefono,
        'fecha_nacimiento': (
            paciente.fecha_nacimiento.isoformat() if paciente.fecha_nacimiento else ''
        ),
        'carnet_igss': paciente.carnet_igss or '',
    })


@login_required
@user_passes_test(es_recepcionista)
def radiologos_por_estudio(request):
    """Usado por el formulario de agendar cita: al elegir el tipo de
    estudio, solo deja elegir entre los radiólogos que realmente lo
    realizan (ej. Celeste solo hace Ultrasonido y Rayos X)."""
    tipo_estudio_id = request.GET.get('tipo_estudio')
    radiologos = Usuario.objects.filter(
        rol=Usuario.ROL_MEDICO_RADIOLOGO, is_active=True, tipos_estudio_asignados__id=tipo_estudio_id,
    ).order_by('username').distinct()
    return JsonResponse({
        'radiologos': [
            {'id': r.id, 'texto': r.get_full_name() or r.username} for r in radiologos
        ],
    })


@login_required
@user_passes_test(es_recepcionista)
def historial_pacientes(request):
    """Listado de pacientes con al menos un estudio ya realizado (informe
    entregado), con búsqueda por nombre/apellido o DPI y filtros por
    convenio, fecha y tipo de estudio. Los que todavía tienen datos
    pendientes (sexo/teléfono/fecha de nacimiento) van primero, con un
    botón para completarlos; a los demás se les muestra de qué
    convenio(s) son sus estudios (COEX, Privado, Emergencia IGSS)."""
    busqueda = (request.GET.get('q') or '').strip()
    filtro_convenio = (request.GET.get('convenio') or '').strip()
    filtro_fecha = (request.GET.get('fecha') or '').strip()
    filtro_tipo_estudio = (request.GET.get('tipo_estudio') or '').strip()

    citas_procesadas = Cita.objects.filter(estado=Cita.ESTADO_PROCESADA)
    if filtro_convenio:
        citas_procesadas = citas_procesadas.filter(convenio=filtro_convenio)
    if filtro_fecha:
        citas_procesadas = citas_procesadas.filter(fecha=filtro_fecha)
    if filtro_tipo_estudio:
        citas_procesadas = citas_procesadas.filter(tipo_estudio_id=filtro_tipo_estudio)

    pacientes_qs = Paciente.objects.filter(
        id__in=citas_procesadas.values('paciente_id')
    ).distinct()
    if busqueda:
        pacientes_qs = pacientes_qs.filter(
            Q(nombre__icontains=busqueda)
            | Q(apellido__icontains=busqueda)
            | Q(dpi__icontains=busqueda)
        )

    pacientes = list(pacientes_qs)

    convenios_por_paciente = {}
    if pacientes:
        convenios_choices = dict(Cita.CONVENIO_CHOICES)
        # .order_by() (sin argumentos) es necesario: si no, el `ordering`
        # por defecto de Cita (fecha, hora) se cuela en el SELECT y hace
        # que .distinct() no junte filas de un mismo convenio con distinta
        # fecha, mostrando el mismo convenio repetido para un paciente.
        filas = (
            Cita.objects.filter(paciente__in=pacientes, estado=Cita.ESTADO_PROCESADA)
            .order_by().values_list('paciente_id', 'convenio').distinct()
        )
        for paciente_id, convenio in filas:
            convenios_por_paciente.setdefault(paciente_id, []).append(
                convenios_choices.get(convenio, convenio)
            )

    for paciente in pacientes:
        paciente.datos_pendientes = paciente.campos_pendientes()
        paciente.convenios_estudios = sorted(convenios_por_paciente.get(paciente.id, []))

    # Los que tienen datos pendientes primero, para que salten a la vista;
    # el resto en orden alfabético, como antes.
    pacientes.sort(key=lambda p: (not p.datos_pendientes, p.nombre, p.apellido))

    contexto = {
        'pacientes': pacientes,
        'busqueda': busqueda,
        'filtro_convenio': filtro_convenio,
        'filtro_fecha': filtro_fecha,
        'filtro_tipo_estudio': filtro_tipo_estudio,
        'tipos_estudio': TipoEstudio.objects.filter(
            id__in=Cita.objects.filter(estado=Cita.ESTADO_PROCESADA).values('tipo_estudio_id')
        ).order_by('nombre'),
    }

    # Búsqueda en vivo: el JS de la página pide solo el listado (sin el
    # HTML completo) a medida que se escribe en el buscador.
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'pacientes/includes/_resultados_pacientes.html', contexto)

    return render(request, 'pacientes/historial_pacientes.html', contexto)


@login_required
@user_passes_test(es_recepcionista)
def historial_paciente(request, paciente_id):
    """Estudios ya realizados (con informe) de un paciente, del más
    reciente al más antiguo."""
    paciente = get_object_or_404(Paciente, id=paciente_id)
    citas = (
        Cita.objects.filter(paciente=paciente, estado=Cita.ESTADO_PROCESADA)
        .select_related('tipo_estudio', 'orden_trabajo')
        .order_by('-fecha', '-hora')
    )
    # Para llenar el combo de "Estudio" del filtro solo con los tipos que
    # este paciente realmente tiene (no el catálogo completo).
    tipos_estudio = sorted({cita.tipo_estudio.nombre for cita in citas})
    return render(request, 'pacientes/historial_paciente.html', {
        'paciente': paciente,
        'citas': citas,
        'tipos_estudio': tipos_estudio,
        'edad': paciente.edad_en(timezone.localdate()),
        'hoy': timezone.localdate(),
    })


@login_required
@user_passes_test(es_recepcionista)
def ver_estudio_historial(request, cita_id):
    """Vista de solo lectura de un estudio ya realizado: imágenes e
    informe, tal como quedaron al terminar el proceso."""
    cita = get_object_or_404(Cita, id=cita_id, estado=Cita.ESTADO_PROCESADA)
    orden = get_object_or_404(OrdenTrabajo, cita=cita)
    return render(request, 'pacientes/ver_estudio_historial.html', {
        'cita': cita,
        'orden': orden,
        'edad': orden.edad_paciente,
        'volver_url': reverse('historial_paciente', args=[cita.paciente_id]),
    })


@login_required
@user_passes_test(es_administrador)
def crear_estudio(request):
    if request.method == 'POST':
        form = CrearTipoEstudioForm(request.POST)
        if form.is_valid():
            tipo_estudio = form.save()
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_CREAR_ESTUDIO,
                descripcion=f'Creó el estudio "{tipo_estudio.nombre}" (precio: {tipo_estudio.precio}).',
            )
            messages.success(request, f'Estudio "{tipo_estudio.nombre}" creado correctamente.')
            return redirect('dashboard')
    else:
        form = CrearTipoEstudioForm()
    return render(request, 'pacientes/crear_estudio.html', {'form': form, 'editando': None})


@login_required
@user_passes_test(es_administrador)
def lista_estudios(request):
    estudios = TipoEstudio.objects.all().order_by('nombre')
    return render(request, 'pacientes/lista_estudios.html', {'estudios': estudios})


@login_required
@user_passes_test(es_administrador)
def editar_estudio(request, estudio_id):
    tipo_estudio = get_object_or_404(TipoEstudio, id=estudio_id)
    if request.method == 'POST':
        form = CrearTipoEstudioForm(request.POST, instance=tipo_estudio)
        if form.is_valid():
            tipo_estudio = form.save()
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_EDITAR_ESTUDIO,
                descripcion=(
                    f'Editó el estudio "{tipo_estudio.nombre}" '
                    f'(precio: {tipo_estudio.precio}, duración: {tipo_estudio.duracion_minutos} min).'
                ),
            )
            messages.success(request, f'Estudio "{tipo_estudio.nombre}" actualizado correctamente.')
            return redirect('lista_estudios')
    else:
        form = CrearTipoEstudioForm(instance=tipo_estudio)
    return render(request, 'pacientes/crear_estudio.html', {'form': form, 'editando': tipo_estudio})


@login_required
@user_passes_test(es_recepcionista)
def seleccionar_horario(request, convenio):
    convenio_nombre = dict(Cita.CONVENIO_CHOICES).get(convenio, convenio)

    reagendar_cita = None
    reagendar_id = request.GET.get('reagendar')
    if reagendar_id:
        reagendar_cita = get_object_or_404(Cita, id=reagendar_id, convenio=convenio)

    hoy = datetime.date.today()
    semana_param = parse_date(request.GET.get('semana', ''))
    inicio = inicio_semana(semana_param or hoy)
    dias = [inicio + datetime.timedelta(days=i) for i in range(len(DIAS_SEMANA))]

    citas_semana = (
        Cita.objects.filter(fecha__gte=dias[0], fecha__lte=dias[-1])
        .exclude(estado=Cita.ESTADO_RECHAZADA)
        .select_related('tipo_estudio')
    )
    if reagendar_cita:
        citas_semana = citas_semana.exclude(id=reagendar_cita.id)

    ocupados_por_dia = {}
    asignados_por_dia = {}
    for cita in citas_semana:
        rango = rango_ocupado_por(cita.fecha, cita.hora, cita.tipo_estudio.duracion_minutos)
        ocupados_por_dia.setdefault(cita.fecha, []).append(rango)
        asignados_por_dia.setdefault(cita.fecha, set()).add(cita.hora)

    filas = [
        {
            'hora': hora,
            'celdas': [
                {
                    'dia': dia,
                    'hora': hora,
                    'pasado': en_el_pasado(dia, hora),
                    'fuera_rango': fuera_de_ventana(dia),
                    'asignado': hora in asignados_por_dia.get(dia, set()),
                    'ocupado': (
                        hora not in asignados_por_dia.get(dia, set())
                        and any(
                            se_cruzan(rango_ocupado_por(dia, hora, PASO_MINUTOS), ocupado)
                            for ocupado in ocupados_por_dia.get(dia, [])
                        )
                    ),
                }
                for dia in dias
            ],
        }
        for hora in horarios_disponibles()
    ]

    contexto = {
        'convenio': convenio,
        'convenio_nombre': convenio_nombre,
        'agendar_url_name': f'agendar_cita_{convenio}',
        'dias': list(zip(DIAS_SEMANA, dias)),
        'filas': filas,
        'semana_anterior': inicio - datetime.timedelta(days=7),
        'mostrar_semana_anterior': inicio > inicio_semana(hoy),
        'semana_siguiente': inicio + datetime.timedelta(days=7),
        'mostrar_semana_siguiente': not fuera_de_ventana(inicio + datetime.timedelta(days=7)),
        'reagendar_cita': reagendar_cita,
        'reagendar_url_name': f'confirmar_reagenda_{convenio}' if reagendar_cita else None,
        'procesar_url_name': f'procesar_citas_{convenio}',
        'maximo_emergencias_por_dia': MAXIMO_EMERGENCIAS_POR_DIA,
    }
    return render(request, 'pacientes/calendario.html', contexto)


def _hay_conflicto_horario(fecha_dt, hora_time, duracion_minutos):
    """¿El horario dado se cruza con alguna cita ya existente ese día?
    (no cuenta las citas rechazadas)."""
    if not fecha_dt or not hora_time:
        return False
    ocupados = [
        rango_ocupado_por(c.fecha, c.hora, c.tipo_estudio.duracion_minutos)
        for c in Cita.objects.filter(fecha=fecha_dt)
        .exclude(estado=Cita.ESTADO_RECHAZADA)
        .select_related('tipo_estudio')
    ]
    rango_nuevo = rango_ocupado_por(fecha_dt, hora_time, duracion_minutos)
    return any(se_cruzan(rango_nuevo, ocupado) for ocupado in ocupados)


@login_required
@user_passes_test(es_recepcionista)
def agendar_cita(request, convenio):
    convenio_nombre = dict(Cita.CONVENIO_CHOICES).get(convenio, convenio)
    calendario_url = reverse(f'calendario_{convenio}')

    datos = request.POST if request.method == 'POST' else request.GET
    fecha = datos.get('fecha')
    hora = datos.get('hora')
    if not fecha or not hora:
        return redirect(calendario_url)

    if request.method == 'POST':
        form = AgendarCitaForm(request.POST, convenio=convenio)
    else:
        form = AgendarCitaForm(initial={'fecha': fecha, 'hora': hora}, convenio=convenio)
    form.fields['fecha'].widget = forms.HiddenInput()
    form.fields['hora'].widget = forms.HiddenInput()

    # Aviso de si el horario elegido ya tiene otra cita encima: se calcula
    # desde ya (sin esperar a elegir el tipo de estudio) usando PASO_MINUTOS
    # como referencia, para que la advertencia de emergencia aparezca apenas
    # se carga la pantalla. Al enviar el formulario se vuelve a calcular con
    # la duración real del estudio elegido, que es la que manda.
    fecha_dt = fecha if isinstance(fecha, datetime.date) else parse_date(fecha)
    try:
        hora_time = hora if isinstance(hora, datetime.time) else datetime.datetime.strptime(hora, '%H:%M').time()
    except (TypeError, ValueError):
        hora_time = None
    hay_conflicto = _hay_conflicto_horario(fecha_dt, hora_time, PASO_MINUTOS)

    if request.method == 'POST' and form.is_valid():
        cd = form.cleaned_data
        if en_el_pasado(cd['fecha'], cd['hora']):
            messages.error(request, 'No se pueden agendar citas en un horario que ya pasó.')
            return redirect(calendario_url)
        if fuera_de_ventana(cd['fecha']):
            messages.error(request, 'Solo se pueden agendar citas hasta 3 semanas después de hoy.')
            return redirect(calendario_url)

        hay_conflicto = _hay_conflicto_horario(cd['fecha'], cd['hora'], cd['tipo_estudio'].duracion_minutos)

        if hay_conflicto and not cd['es_emergencia']:
            form.add_error(
                None,
                'Este horario ya está ocupado por otra cita. Si es una emergencia que debe '
                'agendarse sí o sí en este horario, marcá la casilla de confirmación de '
                'emergencia (más abajo) y volvé a enviar.',
            )
        else:
            if hay_conflicto:
                emergencias_hoy = Cita.objects.filter(
                    fecha=cd['fecha'], es_emergencia_forzada=True,
                ).exclude(estado=Cita.ESTADO_RECHAZADA).count()
                if emergencias_hoy >= MAXIMO_EMERGENCIAS_POR_DIA:
                    messages.error(
                        request,
                        f'Ya se agendaron {MAXIMO_EMERGENCIAS_POR_DIA} citas de emergencia para el '
                        f'{cd["fecha"]}, el máximo permitido por día. Elegí otra fecha.',
                    )
                    return redirect(calendario_url)

            paciente = obtener_o_actualizar_paciente(cd)
            cita = Cita.objects.create(
                paciente=paciente,
                tipo_estudio=cd['tipo_estudio'],
                radiologo=cd['radiologo'],
                convenio=convenio,
                estado=Cita.ESTADO_PENDIENTE,
                fecha=cd['fecha'],
                hora=cd['hora'],
                medico_referente=cd['medico_referente'],
                fecha_sugerida=cd['fecha'],
                hora_sugerida=cd['hora'],
                notas=cd['notas'],
                creada_por=request.user,
                es_emergencia_forzada=hay_conflicto,
            )
            _notificar_cita_asignada(cita)
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_SOLICITAR_CITA,
                descripcion=(
                    ('[EMERGENCIA] ' if hay_conflicto else '')
                    + f'Registró al paciente {paciente.nombre} {paciente.apellido} (DPI {paciente.dpi}) '
                    f'y solicitó cita de {cd["tipo_estudio"]} para {cd["fecha"]} {cd["hora"]} ({convenio}), '
                    f'asignada a {cd["radiologo"]}.'
                    + (
                        ' Se agendó encima de otra cita ya existente por tratarse de una emergencia.'
                        if hay_conflicto else ''
                    )
                ),
            )
            if hay_conflicto:
                messages.success(
                    request,
                    f'Cita de EMERGENCIA agendada para {paciente.nombre} {paciente.apellido} el '
                    f'{cd["fecha"]} a las {cd["hora"]}, encima de otra cita que ya ocupaba ese horario. '
                    f'Se avisó a {cd["radiologo"]} que es una emergencia.',
                )
            else:
                messages.success(
                    request,
                    f'Solicitud enviada a {cd["radiologo"]} para {paciente.nombre} {paciente.apellido} '
                    f'(sugerido: {cd["fecha"]} a las {cd["hora"]}). '
                    'Quedará agendada cuando la radióloga la confirme.',
                )
            return redirect('dashboard')

    return render(request, 'pacientes/agendar_cita.html', {
        'form': form,
        'convenio': convenio,
        'convenio_nombre': convenio_nombre,
        'calendario_url': calendario_url,
        'fecha_valor': fecha,
        'hora_valor': hora,
        'requiere_carnet_igss': convenio in (Cita.CONVENIO_COEX, Cita.CONVENIO_EMERGENCIA_IGSS),
        'hay_conflicto': hay_conflicto,
    })


@login_required
@user_passes_test(es_recepcionista)
def procesar_citas(request, convenio):
    convenio_nombre = dict(Cita.CONVENIO_CHOICES).get(convenio, convenio)

    fecha = parse_date(request.GET.get('fecha', '')) or datetime.date.today()
    citas = (
        Cita.objects.filter(convenio=convenio, fecha=fecha)
        .exclude(estado=Cita.ESTADO_PENDIENTE)
        .select_related('paciente', 'tipo_estudio')
        .order_by('hora')
    )

    return render(request, 'pacientes/procesar_citas.html', {
        'convenio': convenio,
        'convenio_nombre': convenio_nombre,
        'fecha': fecha,
        'hoy': datetime.date.today(),
        'dia_anterior': fecha - datetime.timedelta(days=1),
        'dia_siguiente': fecha + datetime.timedelta(days=1),
        'citas': citas,
        'calendario_url_name': f'calendario_{convenio}',
        'marcar_llegada_url_name': f'marcar_llegada_{convenio}',
        'generar_orden_url_name': f'generar_orden_{convenio}',
        'marcar_ausente_url_name': f'marcar_ausente_{convenio}',
    })


@login_required
@user_passes_test(es_recepcionista)
def marcar_llegada(request, convenio, cita_id):
    cita = get_object_or_404(Cita, id=cita_id, convenio=convenio)
    if request.method == 'POST' and cita.estado == Cita.ESTADO_AGENDADA:
        cita.hora_llegada = timezone.now()
        cita.save(update_fields=['hora_llegada'])
        Bitacora.registrar(
            request=request,
            usuario=request.user,
            accion=Bitacora.ACCION_MARCAR_LLEGADA,
            descripcion=f'Marcó la llegada de {cita.paciente} (cita #{cita.id}).',
        )
        messages.success(request, f'Se registró la llegada de {cita.paciente}.')
    return redirect(f'{reverse(f"procesar_citas_{convenio}")}?fecha={cita.fecha}')


@login_required
@user_passes_test(es_recepcionista)
def generar_orden(request, convenio, cita_id):
    cita = get_object_or_404(Cita, id=cita_id, convenio=convenio)
    volver_url = f'{reverse(f"procesar_citas_{convenio}")}?fecha={cita.fecha}'

    if not cita.hora_llegada:
        messages.error(request, 'Primero hay que marcar la llegada del paciente.')
        return redirect(volver_url)
    if cita.estado != Cita.ESTADO_AGENDADA:
        messages.error(request, 'Esta cita ya no está pendiente de procesar.')
        return redirect(volver_url)

    if request.method == 'POST':
        form = GenerarOrdenForm(request.POST)
        if form.is_valid():
            OrdenTrabajo.objects.create(
                cita=cita,
                motivo=form.cleaned_data['motivo'],
                creada_por=request.user,
            )
            cita.estado = Cita.ESTADO_EN_PROCESO
            cita.save(update_fields=['estado'])
            _notificar_orden_pendiente(cita)
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_GENERAR_ORDEN,
                descripcion=f'Generó la orden de trabajo para {cita.paciente} (cita #{cita.id}).',
            )
            messages.success(
                request, f'Orden de trabajo generada y enviada al técnico para {cita.paciente}.'
            )
            return redirect(volver_url)
    else:
        form = GenerarOrdenForm()

    return render(request, 'pacientes/generar_orden.html', {
        'form': form,
        'cita': cita,
        'edad': cita.paciente.edad_en(cita.fecha),
        'volver_url': volver_url,
    })


@login_required
@user_passes_test(es_tecnico)
def ordenes_pendientes(request):
    ordenes = (
        OrdenTrabajo.objects.filter(cita__estado=Cita.ESTADO_EN_PROCESO)
        .exclude(imagenes__isnull=False)
        .select_related('cita', 'cita__paciente', 'cita__tipo_estudio')
        .distinct()
        .order_by('creada_en')
    )
    return render(request, 'pacientes/ordenes_pendientes.html', {'ordenes': ordenes})


@login_required
@user_passes_test(es_tecnico)
def adjuntar_imagenes(request, orden_id):
    orden = get_object_or_404(OrdenTrabajo, id=orden_id, cita__estado=Cita.ESTADO_EN_PROCESO)
    volver_url = reverse('ordenes_pendientes')

    if request.method == 'POST':
        form = AdjuntarImagenesForm(request.POST, request.FILES)
        if form.is_valid():
            archivos = form.cleaned_data['imagenes']
            for archivo in archivos:
                ImagenEstudio.objects.create(orden=orden, archivo=archivo, subida_por=request.user)
            _notificar_estudio_listo_para_informar(orden.cita)
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_ADJUNTAR_IMAGENES,
                descripcion=(
                    f'Adjuntó {len(archivos)} imagen(es) a la orden de {orden.cita.paciente} '
                    f'(orden #{orden.id}).'
                ),
            )
            messages.success(
                request, f'Imágenes adjuntadas para {orden.cita.paciente}. Ya está lista para la radióloga.'
            )
            return redirect(volver_url)
    else:
        form = AdjuntarImagenesForm()

    return render(request, 'pacientes/adjuntar_imagenes.html', {
        'form': form,
        'cita': orden.cita,
        'orden': orden,
        'edad': orden.edad_paciente,
        'volver_url': volver_url,
    })


@login_required
@user_passes_test(es_radiologo)
def citas_procesadas(request):
    ordenes = (
        OrdenTrabajo.objects.filter(cita__estado=Cita.ESTADO_EN_PROCESO, imagenes__isnull=False)
        .select_related('cita', 'cita__paciente', 'cita__tipo_estudio')
        .distinct()
        .order_by('creada_en')
    )
    return render(request, 'pacientes/citas_procesadas.html', {'ordenes': ordenes})


@login_required
@user_passes_test(es_radiologo)
def adjuntar_informe(request, cita_id):
    cita = get_object_or_404(Cita, id=cita_id, estado=Cita.ESTADO_EN_PROCESO)
    orden = get_object_or_404(OrdenTrabajo, cita=cita)
    if not orden.tiene_imagenes:
        messages.error(request, 'El técnico todavía no adjunta las imágenes de este estudio.')
        return redirect('citas_procesadas')
    volver_url = reverse('citas_procesadas')

    if request.method == 'POST':
        form = AdjuntarInformeForm(request.POST, request.FILES)
        if form.is_valid():
            orden.informe_texto = form.cleaned_data['informe_texto']
            if form.cleaned_data['informe_archivo']:
                orden.informe_archivo = form.cleaned_data['informe_archivo']
            orden.informe_creado_por = request.user
            orden.informe_creado_en = timezone.now()
            orden.save(update_fields=[
                'informe_texto', 'informe_archivo', 'informe_creado_por', 'informe_creado_en',
            ])
            cita.estado = Cita.ESTADO_PROCESADA
            cita.save(update_fields=['estado'])
            _notificar_estudio_completado(cita)
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_ADJUNTAR_INFORME,
                descripcion=f'Adjuntó el informe de {cita.paciente} (cita #{cita.id}).',
            )
            messages.success(request, f'Informe adjuntado para {cita.paciente}.')
            return redirect(volver_url)
    else:
        form = AdjuntarInformeForm()

    return render(request, 'pacientes/adjuntar_informe.html', {
        'form': form,
        'cita': cita,
        'orden': orden,
        'edad': cita.paciente.edad_en(cita.fecha),
        'volver_url': volver_url,
    })


@login_required
@user_passes_test(es_radiologo)
def solicitudes_pendientes(request):
    citas = (
        Cita.objects.filter(estado=Cita.ESTADO_PENDIENTE, radiologo=request.user)
        .select_related('paciente', 'tipo_estudio')
        .order_by('fecha_sugerida', 'hora_sugerida')
    )
    return render(request, 'pacientes/solicitudes_pendientes.html', {'citas': citas})


@login_required
@user_passes_test(es_radiologo)
def revisar_solicitud(request, cita_id):
    cita = get_object_or_404(
        Cita, id=cita_id, estado=Cita.ESTADO_PENDIENTE, radiologo=request.user
    )
    volver_url = reverse('solicitudes_pendientes')
    limite = cita.creada_en.date() + datetime.timedelta(days=LIMITE_DIAS_ADELANTE)

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'rechazar':
            cita.estado = Cita.ESTADO_RECHAZADA
            cita.motivo_rechazo = request.POST.get('motivo_rechazo', '').strip()
            cita.revisada_por = request.user
            cita.revisada_en = timezone.now()
            cita.save(update_fields=['estado', 'motivo_rechazo', 'revisada_por', 'revisada_en'])
            _notificar_cita_rechazada(cita)
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_RECHAZAR_CITA,
                descripcion=(
                    f'Rechazó la solicitud de cita de {cita.paciente} (cita #{cita.id}). '
                    f'Motivo: {cita.motivo_rechazo or "—"}'
                ),
            )
            messages.success(request, f'Solicitud de {cita.paciente} rechazada.')
            return redirect(volver_url)

        fecha = parse_date(request.POST.get('fecha', ''))
        hora_str = request.POST.get('hora', '')
        try:
            hora = datetime.datetime.strptime(hora_str, '%H:%M').time()
        except ValueError:
            hora = None

        if not fecha or not hora:
            messages.error(request, 'Fecha u hora inválida.')
            return redirect(request.path)
        if en_el_pasado(fecha, hora):
            messages.error(request, 'No se puede confirmar una cita en un horario que ya pasó.')
            return redirect(request.path)
        if fecha > limite:
            messages.error(
                request,
                f'La cita debe quedar confirmada antes del {limite} (3 semanas desde que se solicitó).',
            )
            return redirect(request.path)

        cita.fecha = fecha
        cita.hora = hora
        cita.estado = Cita.ESTADO_AGENDADA
        cita.revisada_por = request.user
        cita.revisada_en = timezone.now()
        cita.save(update_fields=['fecha', 'hora', 'estado', 'revisada_por', 'revisada_en'])
        ReporteDiario.objects.get_or_create(fecha=cita.fecha, convenio=cita.convenio)
        _notificar_cita_confirmada(cita)
        Bitacora.registrar(
            request=request,
            usuario=request.user,
            accion=Bitacora.ACCION_CONFIRMAR_CITA,
            descripcion=(
                f'Confirmó la cita de {cita.paciente} para el {fecha} '
                f'a las {hora_str} (cita #{cita.id}).'
            ),
        )
        messages.success(request, f'Cita de {cita.paciente} confirmada para el {fecha} a las {hora_str}.')
        return redirect(volver_url)

    return render(request, 'pacientes/revisar_solicitud.html', {
        'cita': cita,
        'edad': cita.paciente.edad_en(cita.fecha_sugerida or cita.fecha),
        'limite': limite,
        'volver_url': volver_url,
    })


@login_required
@user_passes_test(es_recepcionista)
def marcar_ausente(request, convenio, cita_id):
    cita = get_object_or_404(Cita, id=cita_id, convenio=convenio)
    if request.method == 'POST':
        if cita.fecha > datetime.date.today():
            messages.error(request, 'No se puede marcar ausente una cita antes de la fecha en que le toca.')
        else:
            cita.estado = Cita.ESTADO_AUSENTE
            cita.save(update_fields=['estado'])
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_MARCAR_AUSENTE,
                descripcion=f'Marcó como ausente a {cita.paciente} (cita #{cita.id}).',
            )
            messages.success(request, f'Cita de {cita.paciente} marcada como ausente.')
    return redirect(f'{reverse(f"procesar_citas_{convenio}")}?fecha={cita.fecha}')


@login_required
@user_passes_test(es_recepcionista)
def confirmar_reagenda(request, convenio, cita_id):
    cita = get_object_or_404(Cita, id=cita_id, convenio=convenio)
    calendario_url = reverse(f'calendario_{convenio}')

    if cita.estado != Cita.ESTADO_AUSENTE:
        messages.error(request, 'Solo se pueden reagendar citas marcadas como ausente.')
        return redirect(f'{reverse(f"procesar_citas_{convenio}")}?fecha={cita.fecha}')

    datos = request.POST if request.method == 'POST' else request.GET
    fecha = parse_date(datos.get('fecha', ''))
    hora = datos.get('hora')
    if not fecha or not hora:
        return redirect(f'{calendario_url}?reagendar={cita.id}')

    if request.method == 'POST':
        hora_valor = datetime.datetime.strptime(hora, '%H:%M').time()
        if en_el_pasado(fecha, hora_valor):
            messages.error(request, 'No se pueden reagendar citas a un horario que ya pasó.')
            return redirect(f'{calendario_url}?reagendar={cita.id}')
        if fuera_de_ventana(fecha):
            messages.error(request, 'Solo se pueden reagendar citas hasta 3 semanas después de hoy.')
            return redirect(f'{calendario_url}?reagendar={cita.id}')

        ocupados = [
            rango_ocupado_por(c.fecha, c.hora, c.tipo_estudio.duracion_minutos)
            for c in Cita.objects.filter(fecha=fecha)
            .exclude(estado=Cita.ESTADO_RECHAZADA)
            .exclude(id=cita.id)
            .select_related('tipo_estudio')
        ]
        rango_nuevo = rango_ocupado_por(fecha, hora_valor, cita.tipo_estudio.duracion_minutos)
        if any(se_cruzan(rango_nuevo, ocupado) for ocupado in ocupados):
            messages.error(request, 'Ese horario ya no está disponible: se cruza con otra cita.')
            return redirect(f'{calendario_url}?reagendar={cita.id}')

        cita.fecha = fecha
        cita.hora = hora_valor
        cita.estado = Cita.ESTADO_AGENDADA
        cita.save(update_fields=['fecha', 'hora', 'estado'])
        ReporteDiario.objects.get_or_create(fecha=cita.fecha, convenio=cita.convenio)
        Bitacora.registrar(
            request=request,
            usuario=request.user,
            accion=Bitacora.ACCION_REAGENDAR_CITA,
            descripcion=(
                f'Reagendó la cita de {cita.paciente} para el {fecha} '
                f'a las {hora} (cita #{cita.id}).'
            ),
        )
        messages.success(request, f'Cita de {cita.paciente} reagendada para el {fecha} a las {hora}.')
        return redirect(f'{reverse(f"procesar_citas_{convenio}")}?fecha={fecha}')

    return render(request, 'pacientes/reagendar_confirmar.html', {
        'cita': cita,
        'nueva_fecha': fecha,
        'nueva_hora': hora,
        'calendario_url': f'{calendario_url}?reagendar={cita.id}',
    })


# Registrar Ticket: check-in de pacientes que llegan a Emergencia IGSS sin
# cita agendada. Genera un turno numerado (ver Ticket.save) para la fila de
# atención.
@login_required
@user_passes_test(es_recepcionista)
def registrar_ticket_emergencia(request):
    volver_url = reverse('pantalla_placeholder', kwargs={'clave': 'emergencia_igss'})

    if request.method == 'POST':
        form = RegistrarTicketForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            paciente = obtener_o_actualizar_paciente(cd)
            ticket = Ticket.objects.create(
                paciente=paciente,
                servicio=Ticket.SERVICIO_EMERGENCIA_IGSS,
                prioridad=int(cd['prioridad']),
                motivo=cd['motivo'],
                registrado_por=request.user,
            )
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_REGISTRAR_TICKET,
                descripcion=(
                    f'Registró el ticket {ticket.turno} de Emergencia IGSS para '
                    f'{paciente.nombre} {paciente.apellido} (DPI {paciente.dpi}).'
                ),
            )
            messages.success(
                request, f'Ticket {ticket.turno} registrado para {paciente.nombre} {paciente.apellido}.'
            )
            return redirect('pantalla_turnos_emergencia')
    else:
        form = RegistrarTicketForm()

    return render(request, 'pacientes/registrar_ticket_emergencia.html', {
        'form': form,
        'volver_url': volver_url,
    })


@login_required
@user_passes_test(es_recepcionista)
def pantalla_turnos_emergencia(request):
    cola = (
        Ticket.objects.filter(servicio=Ticket.SERVICIO_EMERGENCIA_IGSS)
        .exclude(estado__in=[Ticket.ESTADO_ATENDIDO, Ticket.ESTADO_AUSENTE])
        .select_related('paciente')
        .order_by('-prioridad', 'creado_en')
    )
    return render(request, 'pacientes/pantalla_turnos_emergencia.html', {
        'cola': cola,
        'siguiente': cola.first(),
        'volver_url': reverse('pantalla_placeholder', kwargs={'clave': 'emergencia_igss'}),
    })


@login_required
@user_passes_test(es_recepcionista)
def procesar_ticket_emergencia(request, ticket_id):
    """Convierte el ticket en una cita EN_PROCESO + orden de trabajo, lista
    para que el técnico la vea en 'Órdenes pendientes'. Se salta agendado y
    revisión del radiólogo porque el paciente ya está en la clínica."""
    ticket = get_object_or_404(Ticket, id=ticket_id, servicio=Ticket.SERVICIO_EMERGENCIA_IGSS)
    volver_url = reverse('pantalla_turnos_emergencia')

    if ticket.estado != Ticket.ESTADO_EN_ESPERA:
        messages.error(request, f'El ticket {ticket.turno} ya fue procesado.')
        return redirect(volver_url)

    if request.method == 'POST':
        form = ProcesarTicketForm(request.POST)
        if form.is_valid():
            ahora = timezone.localtime()
            cita = Cita.objects.create(
                paciente=ticket.paciente,
                tipo_estudio=form.cleaned_data['tipo_estudio'],
                convenio=Cita.CONVENIO_EMERGENCIA_IGSS,
                estado=Cita.ESTADO_EN_PROCESO,
                fecha=ahora.date(),
                hora=ahora.time(),
                hora_llegada=ticket.creado_en,
                notas=ticket.motivo,
                creada_por=request.user,
            )
            OrdenTrabajo.objects.create(
                cita=cita,
                motivo=form.cleaned_data['motivo'],
                creada_por=request.user,
            )
            ticket.estado = Ticket.ESTADO_ATENDIDO
            ticket.atendido_en = timezone.now()
            ticket.cita = cita
            ticket.save(update_fields=['estado', 'atendido_en', 'cita'])
            _notificar_orden_pendiente(cita)
            Bitacora.registrar(
                request=request,
                usuario=request.user,
                accion=Bitacora.ACCION_PROCESAR_TICKET,
                descripcion=(
                    f'Procesó el ticket {ticket.turno} de Emergencia IGSS y generó la orden de trabajo '
                    f'para {ticket.paciente} (cita #{cita.id}).'
                ),
            )
            messages.success(request, f'Ticket {ticket.turno} procesado: la orden ya está con el técnico.')
            return redirect(volver_url)
    else:
        form = ProcesarTicketForm(initial={'motivo': ticket.motivo})

    return render(request, 'pacientes/procesar_ticket_emergencia.html', {
        'form': form,
        'ticket': ticket,
        'volver_url': volver_url,
    })


MAX_NOTIFICACIONES_EN_CAMPANITA = 20


@login_required
def notificaciones_pendientes(request):
    """Endpoint que la campanita de notificaciones consulta cada cierto
    tiempo (ver includes/notificaciones.html) para saber si hay avisos
    nuevos y hacer sonar el aviso."""
    pendientes = request.user.notificaciones.filter(leida=False)
    notificaciones = list(pendientes[:MAX_NOTIFICACIONES_EN_CAMPANITA])
    return JsonResponse({
        'no_leidas': pendientes.count(),
        'notificaciones': [
            {
                'id': n.id,
                'tipo': n.tipo,
                'mensaje': n.mensaje,
                'url': n.url,
                'creada_en': timezone.localtime(n.creada_en).strftime('%d/%m %H:%M'),
            }
            for n in notificaciones
        ],
    })


@login_required
def marcar_notificacion_leida(request, notificacion_id):
    if request.method == 'POST':
        # Los avisos de "datos de paciente pendientes" no se pueden cerrar a
        # mano: se apagan solos cuando el dato realmente se completa (ver
        # completar_datos_paciente). Se valida también aquí, no solo en el
        # JS, para que no se puedan cerrar llamando el endpoint directo.
        request.user.notificaciones.filter(id=notificacion_id).exclude(
            tipo=Notificacion.TIPO_DATOS_PACIENTE_PENDIENTES,
        ).update(leida=True)
    return JsonResponse({'ok': True})


@login_required
def marcar_notificaciones_leidas(request):
    if request.method == 'POST':
        request.user.notificaciones.filter(leida=False).exclude(
            tipo=Notificacion.TIPO_DATOS_PACIENTE_PENDIENTES,
        ).update(leida=True)
    return JsonResponse({'ok': True})


# --- Reportes diarios ---------------------------------------------------
#
# Un ReporteDiario existe por (convenio, fecha). Se crea automáticamente en
# estado "borrador" la primera vez que un radiólogo confirma una cita de ese
# convenio para esa fecha (ver revisar_solicitud). Su contenido (las citas)
# no se guarda aparte: se calcula en el momento a partir de Cita, así nunca
# queda desactualizado. La recepcionista puede editar el médico referente de
# cada fila mientras el reporte esté en borrador, y enviarlo cuando termine
# el día; el administrador financiero solo ve los reportes ya enviados, y
# puede descargarlos en PDF o Excel.

DIAS_LARGOS_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
MESES_LARGOS_ES = [
    'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
    'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
]

COLUMNAS_REPORTE = [
    'No.', 'Hora', 'Nombre del Paciente', 'Edad', 'Estudio',
    'Técnico', 'Médico Referente', 'Emerg', 'Radiólogo', 'Precio',
]


def _fecha_larga_es(fecha):
    dia_semana = DIAS_LARGOS_ES[fecha.weekday()]
    mes = MESES_LARGOS_ES[fecha.month - 1]
    return f'{dia_semana} {fecha.day} de {mes} de {fecha.year}'


def _validar_convenio(convenio):
    if convenio not in dict(Cita.CONVENIO_CHOICES):
        raise Http404('Convenio no válido.')


def _solo_ve_reportes_enviados(user):
    """True si debe ver los reportes en modo solo-lectura (solo los ya
    enviados): administrador financiero o administrador general, distinto
    de una recepcionista que también tenga alguno de esos roles de prueba
    (superusuario) — a esa se le sigue tratando como recepcionista."""
    return (
        (es_administrador_financiero(user) or es_administrador(user))
        and not es_recepcionista(user)
    )


def _filas_reporte(reporte):
    filas = []
    for indice, cita in enumerate(reporte.citas(), start=1):
        tecnico = cita.tecnico_asignado
        filas.append({
            'no': indice,
            'cita_id': cita.id,
            'hora': cita.hora.strftime('%H:%M'),
            'paciente': f'{cita.paciente.nombre} {cita.paciente.apellido}',
            'edad': cita.paciente.edad_en(cita.fecha),
            'estudio': cita.tipo_estudio.nombre,
            'tecnico': (tecnico.get_full_name() or tecnico.username) if tecnico else '',
            'medico_referente': cita.medico_referente,
            'emerg': 'X' if cita.convenio == Cita.CONVENIO_EMERGENCIA_IGSS else '',
            'radiologo': (
                (cita.radiologo.get_full_name() or cita.radiologo.username) if cita.radiologo else ''
            ),
            'precio': cita.tipo_estudio.precio,
            'ausente': cita.estado == Cita.ESTADO_AUSENTE,
        })
    return filas


def _pacientes_con_datos_pendientes(reporte):
    """Pacientes del reporte (sexo/teléfono/fecha de nacimiento) que
    todavía tienen algún dato pendiente de llenar. Mientras existan, el
    reporte no se puede enviar al administrador."""
    pendientes = {}
    for cita in reporte.citas():
        paciente = cita.paciente
        if paciente.id in pendientes:
            continue
        campos = paciente.campos_pendientes()
        if campos:
            pendientes[paciente.id] = {
                'nombre': f'{paciente.nombre} {paciente.apellido}',
                'campos': campos,
            }
    return sorted(pendientes.values(), key=lambda p: p['nombre'])


def _reporte_pdf_bytes(reporte, filas):
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        topMargin=1 * cm, bottomMargin=1 * cm, leftMargin=1 * cm, rightMargin=1 * cm,
    )
    estilos = getSampleStyleSheet()
    titulo = Paragraph(
        f'INFORME DIARIO "{reporte.get_convenio_display().upper()}"<br/>'
        f'FECHA: {_fecha_larga_es(reporte.fecha).upper()}',
        estilos['Title'],
    )

    datos = [COLUMNAS_REPORTE]
    for fila in filas:
        datos.append([
            fila['no'], fila['hora'], fila['paciente'], fila['edad'] if fila['edad'] is not None else '',
            fila['estudio'], fila['tecnico'], fila['medico_referente'], fila['emerg'], fila['radiologo'],
            f"Q{fila['precio']:.2f}",
        ])
    datos.append(['', '', '', '', '', '', '', '', 'TOTAL DEL DÍA', f'Q{reporte.total():.2f}'])

    tabla = Table(datos, repeatRows=1)
    ultima_fila = len(datos) - 1
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c4b5fd')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN', (0, ultima_fila), (-3, ultima_fila)),
        ('BACKGROUND', (-2, ultima_fila), (-1, ultima_fila), colors.HexColor('#fef08a')),
        ('FONTNAME', (-2, ultima_fila), (-1, ultima_fila), 'Helvetica-Bold'),
    ]))
    doc.build([titulo, Spacer(1, 12), tabla])
    return buffer.getvalue()


def _reporte_xlsx_bytes(reporte, filas):
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    ws.merge_cells('A1:J1')
    ws['A1'] = f'INFORME DIARIO "{reporte.get_convenio_display().upper()}"'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f'FECHA: {_fecha_larga_es(reporte.fecha).upper()}'
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(COLUMNAS_REPORTE)
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='7C3AED')

    for fila in filas:
        ws.append([
            fila['no'], fila['hora'], fila['paciente'], fila['edad'], fila['estudio'],
            fila['tecnico'], fila['medico_referente'], fila['emerg'], fila['radiologo'],
            float(fila['precio']),
        ])

    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=9, value='TOTAL DEL DÍA').font = Font(bold=True)
    celda_total = ws.cell(row=fila_total, column=10, value=float(reporte.total()))
    celda_total.font = Font(bold=True)
    celda_total.fill = PatternFill('solid', fgColor='FEF08A')

    from openpyxl.utils import get_column_letter

    encabezados_fila = 4
    for indice_columna in range(1, len(COLUMNAS_REPORTE) + 1):
        letra = get_column_letter(indice_columna)
        longitud = max(
            (
                len(str(ws.cell(row=fila, column=indice_columna).value))
                for fila in range(encabezados_fila, ws.max_row + 1)
                if ws.cell(row=fila, column=indice_columna).value is not None
            ),
            default=10,
        )
        ws.column_dimensions[letra].width = longitud + 2

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@login_required
@user_passes_test(puede_ver_reportes_diarios)
def lista_reportes_diarios(request, convenio):
    convenio_nombre = dict(Cita.CONVENIO_CHOICES).get(convenio, convenio)
    solo_enviados = _solo_ve_reportes_enviados(request.user)

    # Red de seguridad: normalmente el reporte de una fecha se crea en el
    # momento en que se confirma o reagenda una cita para ella, pero por si
    # falta alguno (p. ej. citas ya confirmadas antes de tener esta lógica),
    # se completa aquí cualquier fecha con citas confirmadas que todavía no
    # tenga su ReporteDiario.
    fechas_con_citas = (
        Cita.objects.filter(convenio=convenio)
        .exclude(estado__in=(Cita.ESTADO_PENDIENTE, Cita.ESTADO_RECHAZADA))
        .values_list('fecha', flat=True)
        .distinct()
    )
    fechas_existentes = set(
        ReporteDiario.objects.filter(convenio=convenio).values_list('fecha', flat=True)
    )
    for fecha in fechas_con_citas:
        if fecha not in fechas_existentes:
            ReporteDiario.objects.get_or_create(fecha=fecha, convenio=convenio)

    # Los reportes ya creados son el registro permanente: una vez que
    # existen, deben seguir apareciendo aunque las citas que los originaron
    # cambien después (se reagenden a otra fecha, se rechacen, etc.). Por
    # eso se listan desde ReporteDiario y no recalculando a partir de Cita
    # en cada visita. Solo se muestran días anteriores a hoy: el reporte de
    # un día que todavía no terminó puede cambiar (citas que se cancelan,
    # reagendan o completan durante el día), así que hasta que el día pasa
    # no se puede confiar en el conteo de estudios realizados / cancelados /
    # reagendados / finalizados.
    reportes = ReporteDiario.objects.filter(
        convenio=convenio, fecha__lt=timezone.localdate(),
    ).order_by('-fecha')
    if solo_enviados:
        reportes = reportes.filter(estado=ReporteDiario.ESTADO_ENVIADO)

    return render(request, 'pacientes/lista_reportes_diarios.html', {
        'convenio': convenio,
        'convenio_nombre': convenio_nombre,
        'reportes': reportes,
        'solo_enviados': solo_enviados,
    })


@login_required
@user_passes_test(puede_ver_reportes_diarios)
def ver_reporte_diario(request, convenio, fecha):
    _validar_convenio(convenio)
    fecha_valor = parse_date(fecha)
    if not fecha_valor:
        raise Http404('Fecha no válida.')

    reporte = get_object_or_404(ReporteDiario, convenio=convenio, fecha=fecha_valor)
    solo_lectura = _solo_ve_reportes_enviados(request.user)
    if solo_lectura and reporte.estado != ReporteDiario.ESTADO_ENVIADO:
        raise Http404('Este reporte todavía no fue enviado.')

    editable = not solo_lectura and reporte.estado == ReporteDiario.ESTADO_BORRADOR

    if request.method == 'POST':
        if not editable:
            messages.error(request, 'Este reporte ya no se puede editar.')
            return redirect(request.path)
        for cita in reporte.citas():
            valor = request.POST.get(f'medico_referente_{cita.id}', '').strip()
            if valor != cita.medico_referente:
                cita.medico_referente = valor
                cita.save(update_fields=['medico_referente'])
        messages.success(request, 'Cambios guardados.')
        return redirect(request.path)

    pacientes_pendientes = _pacientes_con_datos_pendientes(reporte)
    return render(request, 'pacientes/reporte_diario.html', {
        'reporte': reporte,
        'convenio_nombre': dict(Cita.CONVENIO_CHOICES).get(convenio, convenio),
        'columnas': COLUMNAS_REPORTE,
        'filas': _filas_reporte(reporte),
        'total': reporte.total(),
        'editable': editable,
        'solo_lectura': solo_lectura,
        'lista_url_name': f'lista_reportes_diarios_{convenio}',
        'pacientes_pendientes': pacientes_pendientes,
    })


@login_required
@user_passes_test(es_recepcionista)
def enviar_reporte_diario(request, convenio, fecha):
    _validar_convenio(convenio)
    fecha_valor = parse_date(fecha)
    if not fecha_valor:
        raise Http404('Fecha no válida.')

    reporte = get_object_or_404(ReporteDiario, convenio=convenio, fecha=fecha_valor)
    volver_url = reverse('ver_reporte_diario', args=[convenio, fecha])

    if request.method == 'POST' and reporte.estado == ReporteDiario.ESTADO_BORRADOR:
        if fecha_valor >= timezone.localdate():
            messages.error(
                request,
                'No se puede enviar el reporte de hoy ni de una fecha futura: todavía puede '
                'haber citas de ese día que se cancelen, reagenden o finalicen. Esperá a que '
                'el día termine.',
            )
            return redirect(volver_url)
        pendientes = _pacientes_con_datos_pendientes(reporte)
        if pendientes:
            nombres = ', '.join(paciente['nombre'] for paciente in pendientes)
            messages.error(
                request,
                'No se puede enviar el reporte: hay pacientes con datos pendientes de llenar '
                f'(sexo, teléfono o fecha de nacimiento): {nombres}.',
            )
            return redirect(volver_url)

        reporte.estado = ReporteDiario.ESTADO_ENVIADO
        reporte.enviado_por = request.user
        reporte.enviado_en = timezone.now()
        reporte.save(update_fields=['estado', 'enviado_por', 'enviado_en'])
        Bitacora.registrar(
            request=request,
            usuario=request.user,
            accion=Bitacora.ACCION_ENVIAR_REPORTE_DIARIO,
            descripcion=(
                f'Envió el reporte diario de {dict(Cita.CONVENIO_CHOICES).get(convenio, convenio)} '
                f'del {fecha_valor} al administrador.'
            ),
        )
        _notificar_reporte_enviado(reporte, request.user)
        messages.success(request, f'Reporte del {fecha_valor} enviado al administrador.')
    return redirect(volver_url)


@login_required
@user_passes_test(puede_descargar_reportes_diarios)
def descargar_reporte_pdf(request, convenio, fecha):
    _validar_convenio(convenio)
    fecha_valor = parse_date(fecha)
    if not fecha_valor:
        raise Http404('Fecha no válida.')
    reporte = get_object_or_404(
        ReporteDiario, convenio=convenio, fecha=fecha_valor, estado=ReporteDiario.ESTADO_ENVIADO,
    )
    contenido = _reporte_pdf_bytes(reporte, _filas_reporte(reporte))
    respuesta = HttpResponse(contenido, content_type='application/pdf')
    respuesta['Content-Disposition'] = f'attachment; filename="reporte_{convenio}_{fecha_valor}.pdf"'
    return respuesta


@login_required
@user_passes_test(puede_descargar_reportes_diarios)
def descargar_reporte_xlsx(request, convenio, fecha):
    _validar_convenio(convenio)
    fecha_valor = parse_date(fecha)
    if not fecha_valor:
        raise Http404('Fecha no válida.')
    reporte = get_object_or_404(
        ReporteDiario, convenio=convenio, fecha=fecha_valor, estado=ReporteDiario.ESTADO_ENVIADO,
    )
    contenido = _reporte_xlsx_bytes(reporte, _filas_reporte(reporte))
    respuesta = HttpResponse(
        contenido, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    respuesta['Content-Disposition'] = f'attachment; filename="reporte_{convenio}_{fecha_valor}.xlsx"'
    return respuesta
